from tkinter import *
import webbrowser
from tkinter import filedialog
from amount_to_million import amount_to_million
from amount_to_crore import amount_to_crore
import openpyxl as xl
from openpyxl.styles import Font
import os

def removingDoubleSpace(value):
    return str(value).replace('  ', ' ')


def callback(url):
    webbrowser.open_new(url)


def uploadExcel():
    filetypes = (
        ('Excel 2010 Files', '*.xlsx'),
    )

    filename = filedialog.askopenfilename(
        title='Open a file',
        filetypes=filetypes)
    textLocation.config(text=filename)

    # file = open("pdf.txt", "w")
    # file.write(filename)
    # file.close()
    # file = open("pdf.txt", "r")
    # data = file.read()
    # if len(data) > 0:
    #     getPDFLocation.config(text=data, foreground="black", font=(8))


def amountTranslator():
    pick = pickText.get()
    paste = pasteText.get()
    location = textLocation.cget("text")
    if len(location) < 1:
        getTextLocation.config(foreground="red")
    else:
        getTextLocation.config(foreground="black")

        currency = unitText.get()
        desktop = location
        wb = xl.load_workbook(desktop, data_only=True)
        # wb = xl.load_workbook(r'C:\Users\HYSTOU\Desktop\sest.xlsx')
        sheet = wb.worksheets[0]
        cell = sheet.cell(1, 1)

        for row in range(1, sheet.max_row + 1):
            try:
                cell = sheet[str(pick)+str(row)]
                if (var.get()) == 1:
                    amount_in_words = amount_to_million(cell.value)
                elif (var.get()) == 2:
                    amount_in_words = amount_to_crore(cell.value)

                amount_in_words_cell = sheet[str(paste)+str(row)]
                if amount_in_words != "Amount is too big to process":
                    amount_in_words_cell.value = removingDoubleSpace(amount_in_words + " " + currency)
                else:
                    amount_in_words_cell.value = amount_in_words
                    amount_in_words_cell = sheet.cell(row, 2)
                    amount_in_words_cell.value == "Amount is too big to process "
                    amount_in_words_cell.font = Font(bold=True, name='Arial', color='FF0000')
            except TypeError:
                amount_in_words_cell = sheet.cell(row, 2)
                amount_in_words_cell.value = "Not Number. VALUE ERROR"
                amount_in_words_cell.font = Font(bold=True, name='Arial', color='FF0000')
            except ValueError:
                amount_in_words_cell = sheet[str(paste)+str(row)]
                amount_in_words_cell.value = "Not Number. VALUE ERROR"
                amount_in_words_cell.font = Font(bold=True, name='Arial', color='FF0000')

        while True:
            try:
                wb.save(desktop)
                os.startfile(location)
                print('done')
                break
            except PermissionError:
                print("")
                input("It seems the file is already open. Close the excel file and Press Enter to Conetinue...")




root = Tk()
root.resizable(0,0)
root.iconbitmap('icon.ico')
root.title('Amount Translator - V-2.0')
root.geometry("542x760")
root.configure(bg="white")

var = IntVar(None, 1)
location = ''


# myFont = font.Font(family='Helvetica', size=20, weight='bold')

img=PhotoImage(file='Amount Translator.PNG')
label = Label(root, image=img)
label.configure(foreground="black")
label.configure(bg="white")
label.place(x=2, y=2)

getTextLocation = Label(root, text="Select Excel File", font=("Comic Sans MS", 25, 'bold'))
getTextLocation.configure(bg="white")
getTextLocation.place(x=18, y=95)

textLocation = Label(root, text="", font=(10))
textLocation.config(bg='white')
textLocation.place(x=20, y=142)

textLabel = Label(root, text="Make Sure that values needed to be translated are in first column of first sheet. Translated Values will be added in second column. All vaues present in second column will be destroyed", font=(6), wraplength=520, justify='left', borderwidth=1, relief="solid", padx=7, pady=7)
textLabel.configure(bg="light blue")
textLabel.configure(bd=2)
textLabel.configure(foreground="purple")
textLabel.place(x=10, y=170)


browseButtonTXT = Button(root, text='Select File', command=uploadExcel)
browseButtonTXT.place(x=425, y=108)

radioButtonLabel = Label(root, text="Select Translation Mode", font=("Comic Sans MS", 25, 'bold'), justify='center')
radioButtonLabel.configure(bg="white")
radioButtonLabel.place(x=10, y=245)

radioButtonMillion = Radiobutton(root, text="Translation into Millions", variable=var,  value=1, font=("Arial Rounded MT Bold", 14))
radioButtonMillion.configure(bg="white")
radioButtonMillion.place(x=10, y=295)

radioButtonCrore = Radiobutton(root, text="Translation into Crore and Lac", variable=var, value=2, font=('Arial Rounded MT Bold', 14))
radioButtonCrore.configure(bg="white")
radioButtonCrore.place(x=10, y=330)

radioColumnLabel = Label(root, text="Select Columns", font=("Comic Sans MS", 25, 'bold'), justify='center')
radioColumnLabel.configure(bg="white")
radioColumnLabel.place(x=10, y=372)

pickLabel = Label(root, text="Pick Values From Column: ", font=('Arial Rounded MT Bold', 14), justify='center')
pickLabel.configure(bg="white")
pickLabel.place(x=10, y=433)

pickText = Entry(root, width=5, textvariable=(StringVar(root, value='A')), foreground='blue', font=("Arial", 12, 'bold'))
pickText.place(x=257, y=438)

pasteLabel = Label(root, text="Paste Values Into Column: ", font=('Arial Rounded MT Bold', 14), justify='center')
pasteLabel.configure(bg="white")
pasteLabel.place(x=10, y=467)

pasteText = Entry(root, width=5, textvariable=(StringVar(root, value='B')), foreground='blue', font=("Arial", 12, 'bold'))
pasteText.place(x=257, y=470)


unitLabel = Label(root, text="Add unit: ", font=("Arial", 12, 'bold'), justify='center')
unitLabel.configure(bg="white")
unitLabel.place(x=10, y=525)

unitText = Entry(root, width=20, foreground='blue', font=("Arial", 12, 'bold'))
unitText.place(x=90, y=526)

unitDescription = Label(root, text="Adding unit will add the value entered right after the translated value. Like Rupees Only or Dollars Only", font=(6), wraplength=520, justify='left', borderwidth=1, relief="solid", padx=7, pady=7)
unitDescription.configure(bg="light blue")
unitDescription.configure(bd=2)
unitDescription.configure(foreground="purple")
unitDescription.place(x=10, y=555)

startButton = Button(root, text="S T A R T", font=("Arial", 15, 'bold'), justify='center', command=amountTranslator)
startButton.configure(foreground="black")
startButton.configure(bg="light green")
startButton.place(x=215, y=670)

footer = Label(root, text="softwares.rubick.org", font=(14), cursor="hand2")
footer.bind("<Button-1>", lambda e: callback("http://softwares.rubick.org"))
footer.configure(foreground="white")
footer.configure(bg="black")
footer.pack(side=BOTTOM)
root.mainloop()