from amount_to_million import amount_to_million
from amount_to_crore import amount_to_crore
import openpyxl as xl
from openpyxl.styles import PatternFill, Font
import os
#desktop = os.path.expanduser("~\Desktop\sest.xlsx")


print("""
Amount To Word Converter - G*CB
===============================

How To Use:
1- Create an excel fie in your desktop (Will only work with file present on desktop) having all the values on column A
(Starting from 1st row) and all the values will be translated into words in column b. Make sure the Sheet Name in excel
is 'Sheet1' (which is default name of excel sheet)
2- Make sure to Save and close the file after adding values in Column A. (If the file is open or unsaved the app will
not work)
2- Run the Amount To Word Translator.exe
3- Enter Currency: Add any currency you want like Rupees Only, Dollars Only The word you add here will be added to the
translations and if you don't want any curreccy to be added just press enter
4- Enter Sheet Name: Make sure to type exact spelling (Case sensitive) with extension Like abc.xlsx (not writing the
wont make the application work
5- Open the file and see the output

Good Luck

Regards,
Usman Khawar
""")





currency = input('Enter Currency (Dollars Only or Rupees Only): ')

while True:
    sheet_name = input('Enter Sheet name with extension (e.g. accounts.xlsx): ')
    if sheet_name.__contains__("xls" or "xlsx") and len(sheet_name) > 5:
        break
    else:
        input(
            "File name is incorrect or contain inappropriate extensions. Only xls and xlsx are acceptable. Press Enter To Continue...")

desktop = os.path.expanduser("~\Desktop\\" + sheet_name)

wb = xl.load_workbook(desktop, data_only=True)
#wb = xl.load_workbook(r'C:\Users\HYSTOU\Desktop\sest.xlsx')
sheet = wb.worksheets[0]
cell = sheet.cell(1, 1)


for row in range(1, sheet.max_row + 1):
    try:
        cell = sheet.cell(row, 1)
        amount_in_words = amount_to_crore(cell.value)
        amount_in_words_cell = sheet.cell(row, 2)
        if amount_in_words != "Amount is too big to process":
            amount_in_words_cell.value = amount_in_words + " " + currency
        else:
            amount_in_words_cell.value = amount_in_words
            amount_in_words_cell = sheet.cell(row, 2)
            amount_in_words_cell.value == "Amount is too big to process "
            amount_in_words_cell.font = Font(bold=True, name='Arial', color='FF0000')
    except TypeError:
        amount_in_words_cell = sheet.cell(row, 2)
        amount_in_words_cell.value = "Not Number. VALUE ERROR"
        amount_in_words_cell.font = Font(bold=True, name='Arial', color='FF0000')
    except ValueError:
        amount_in_words_cell = sheet.cell(row, 2)
        amount_in_words_cell.value = "Not Number. VALUE ERROR"
        amount_in_words_cell.font = Font(bold=True, name='Arial', color='FF0000')

while True:
    try:
        wb.save(desktop)
        print('done')
        break
    except PermissionError:
        print("")
        input("It seems the file is already open. Close the excel file and Press Enter to Conetinue...")














# print(amount_to_words(123456789) + " Rupees Only")
# print(mobile_numbers("I am good"))